import time
import os
import re
from datetime import datetime

# Для YouTube Data API
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# webdriver-manager
from webdriver_manager.chrome import ChromeDriverManager

# openpyxl
from openpyxl import Workbook, load_workbook

# ====== Настройки ======
XLSX_INPUT = "channel_info.xlsx"    # Входной файл
XLSX_OUTPUT = "final_channels.xlsx" # Выходной файл результатов

# Ваш API-ключ для YouTube Data API
DEVELOPER_KEY = "AIzaSyDWpLkA5QbQjqz-pfaV03FslYXPGLOn9zg"

MAX_CHANNELS = None  # Лимит на кол-во обрабатываемых каналов

def iso_to_readable(iso_dt_str: str) -> str:
    """
    Принимает строку в формате ISO-8601, например "2025-03-17T16:00:01Z",
    и возвращает "YYYY-MM-DD HH:MM:SS" (UTC).
    """
    if not iso_dt_str:
        return ""
    try:
        iso_dt_str = iso_dt_str.replace("Z", "+00:00")
        dt = datetime.fromisoformat(iso_dt_str)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except:
        return iso_dt_str


def get_webdriver():
    """
    Настраиваем ChromeDriver.
    """
    chrome_options = Options()
    chrome_options.add_argument("--lang=en-US")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--disable-software-rasterizer")
    # chrome_options.add_argument("--headless")  # При необходимости
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-notifications")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_window_size(1920, 1080)
    return driver


def normalize_channel_url(raw_url: str) -> str:
    """
    Преобразует handle/частичную ссылку в формальную ссылку на канал
    + добавляет ?hl=en&gl=US.
    """
    raw_url = raw_url.strip()
    if raw_url.startswith("http://") or raw_url.startswith("https://"):
        if "hl=en" not in raw_url and "gl=US" not in raw_url:
            if "?" in raw_url:
                raw_url += "&hl=en&gl=US"
            else:
                raw_url += "?hl=en&gl=US"
        return raw_url
    else:
        raw_url = raw_url.lstrip("/")
        full_url = "https://www.youtube.com/" + raw_url
        if "?" in full_url:
            full_url += "&hl=en&gl=US"
        else:
            full_url += "?hl=en&gl=US"
        return full_url


def get_channel_id_from_handle_selenium(handle: str) -> str:
    """
    Открывает https://www.youtube.com/<handle> и пытается извлечь channelId.
    Ищет <link rel="canonical" href=".../channel/UCxxx" /> или "channelId":"UCxxx".
    Возвращает 'UCxxx...' или "", если не найдено.
    """
    driver_local = get_webdriver()
    try:
        handle_str = handle.lstrip("/")
        url = f"https://www.youtube.com/{handle_str}"
        print(f"[LOG] -> Открываем для поиска channelId: {url}")
        driver_local.get(url)
        time.sleep(5)

        page_source = driver_local.page_source

        # 1) <link rel="canonical" href="https://www.youtube.com/channel/UCxxxx"/>
        canon_regex = r'<link\s+rel="canonical"\s+href="https://www\.youtube\.com/channel/(UC[0-9A-Za-z_\-]+)"'
        m1 = re.search(canon_regex, page_source)
        if m1:
            cid = m1.group(1)
            print(f"[LOG] -> Найден channelId через canonical: {cid}")
            return cid

        # 2) "channelId":"UCxxxx"
        chid_regex = r'"channelId":"(UC[0-9A-Za-z_\-]+)"'
        m2 = re.search(chid_regex, page_source)
        if m2:
            cid = m2.group(1)
            print(f"[LOG] -> Найден channelId внутри скрипта: {cid}")
            return cid

        print("[LOG] -> Не удалось найти channelId.")
        return ""

    except Exception as e:
        print(f"[LOG] -> Ошибка при извлечении channelId: {e}")
        return ""
    finally:
        driver_local.quit()


def parse_subscribers_to_int(subs_text):
    """
    Преобразует строку вида "12.3K subscribers" -> 12300 и т.д.
    """
    if not subs_text:
        return None
    text = subs_text.lower().replace("subscribers", "").strip()
    text = text.replace(" ", "").replace(",", "")

    try:
        if text.endswith("k"):
            val = text[:-1]
            return int(float(val) * 1000)
        elif text.endswith("m"):
            val = text[:-1]
            return int(float(val) * 1000000)
        else:
            return int(float(text))
    except:
        return None


def extract_emails_from_text(text):
    """
    Ищем email в тексте (регэксп).
    """
    pattern = r'[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z0-9-.]+'
    return re.findall(pattern, text)


def guess_name_surname(channel_name):
    """
    Наивно считаем, что если название канала в 2 слова - это Имя Фамилия.
    """
    parts = channel_name.split()
    if len(parts) == 2:
        return channel_name
    return ""


def get_city_country_from_about(driver):
    """
    Пытаемся найти "Location: ..." или "Lives in ..." на вкладке About.
    """
    try:
        about_sections = driver.find_elements(By.CSS_SELECTOR, "ytd-channel-about-metadata-renderer div#description-container")
        if about_sections:
            text = about_sections[0].text
            location_pattern = r'(Location[:\s]+[^\n]+|Lives in\s+[^\n]+)'
            match = re.search(location_pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
    except:
        pass
    return ""


def sum_likes_comments_via_api(video_ids, developer_key=DEVELOPER_KEY):
    """
    Суммируем likeCount и commentCount для списка роликов (videoIds) через API.
    """
    if not video_ids:
        return (0, 0)

    youtube = build("youtube", "v3", developerKey=developer_key)
    total_likes = 0
    total_comments = 0

    chunk_size = 50
    for i in range(0, len(video_ids), chunk_size):
        chunk = video_ids[i:i+chunk_size]
        id_str = ",".join(chunk)
        try:
            response = youtube.videos().list(
                part="statistics",
                id=id_str
            ).execute()
        except HttpError as e:
            # Проверяем, не исчерпана ли квота
            if e.resp.status in [403, 429] or 'quotaExceeded' in str(e.content):
                print("[LOG] -> Достигнут лимит квоты YouTube Data API (quotaExceeded). Прерываем обработку.")
                raise  # Пробрасываем исключение вверх для остановки скрипта
            else:
                print(f"[LOG] -> HttpError при получении статистики видео: {e}")
                continue

        items = response.get("items", [])
        for it in items:
            stats = it.get("statistics", {})
            like_str = stats.get("likeCount", "0")
            comm_str = stats.get("commentCount", "0")

            try:
                total_likes += int(like_str)
            except:
                pass
            try:
                total_comments += int(comm_str)
            except:
                pass

    return (total_likes, total_comments)


def get_newest_and_oldest_video_date_in_playlist(playlist_id: str, youtube) -> tuple:
    """
    Ищем самое новое (max) и самое старое (min) videoPublishedAt в плейлисте uploads.
    Возвращаем (newest, oldest) в ISO-формате.
    """
    if not playlist_id:
        return ("", "")

    newest_date = ""
    oldest_date = ""
    next_page_token = None

    while True:
        try:
            resp = youtube.playlistItems().list(
                part="contentDetails",
                playlistId=playlist_id,
                maxResults=50,
                pageToken=next_page_token
            ).execute()
        except HttpError as e:
            # Проверяем, не исчерпана ли квота
            if e.resp.status in [403, 429] or 'quotaExceeded' in str(e.content):
                print("[LOG] -> Достигнут лимит квоты YouTube Data API (quotaExceeded). Прерываем обработку.")
                raise
            else:
                print(f"[LOG] -> HttpError при получении плейлиста: {e}")
                break

        items = resp.get("items", [])
        if not items:
            break

        for item in items:
            vid_pub = item["contentDetails"].get("videoPublishedAt", "")
            if not vid_pub:
                continue
            if newest_date == "" or vid_pub > newest_date:
                newest_date = vid_pub
            if oldest_date == "" or vid_pub < oldest_date:
                oldest_date = vid_pub

        next_page_token = resp.get("nextPageToken")
        if not next_page_token:
            break

    return (newest_date, oldest_date)


def chunked(iterable, n):
    """
    Разбивает iterable на списки по n элементов.
    """
    for i in range(0, len(iterable), n):
        yield iterable[i:i+n]


def parse_duration_to_seconds(duration_iso8601: str) -> int:
    """
    Преобразует ISO 8601 продолжительность (например 'PT4M13S', 'PT59S', 'PT1H2M30S') в кол-во секунд (int).
    """
    pattern = re.compile(
        r'PT'                  # постоянный префикс
        r'(?:(\d+)H)?'         # часы (\d+H) – необязательно
        r'(?:(\d+)M)?'         # минуты (\d+M) – необязательно
        r'(?:(\d+)S)?'         # секунды (\d+S) – необязательно
    )
    match = pattern.match(duration_iso8601)
    if not match:
        return 0

    hours = int(match.group(1) or 0)
    minutes = int(match.group(2) or 0)
    seconds = int(match.group(3) or 0)
    return hours * 3600 + minutes * 60 + seconds


def process_channel(driver, raw_channel_handle):
    """
    Основная логика: 
      1) channelId через Selenium
      2) через API (part="snippet,brandingSettings,topicDetails,contentDetails,statistics") 
         получаем creation_date_api, country, topics, 
         first/last video published, total_views,
         а также плейлист uploads для подсчёта total_videos, num_videos (не shorts), num_shorts
      3) Через Selenium собираем email, кол-во подписчиков, город/страна (About), и т.д.
      4) Суммируем лайки/комментарии (estimated_likes, estimated_comments) через API.
    """
    data = {
        "channel_id": "",
        "channel_name": "",
        "first_last_name": "",
        "city_country": "",
        "email": "",
        "num_subscribers": None,

        "total_videos": 0,  # всего видео (включая shorts)
        "num_videos": 0,    # обычные (не shorts)
        "num_shorts": 0,    # shorts
        "total_views": None,

        "channel_creation_date_api": "",
        "channel_country_api": "",
        "channel_topics_api": "",
        "first_video_published_api": "",
        "last_video_published_api": "",

        # Эти поля заполняем, но не используем в финальном выводе:
        "creation_date": "",
        "first_video_date": "",
        "last_video_date": "",

        "num_following_channels": 0,
        "estimated_likes": 0,
        "estimated_comments": 0
    }

    print(f"[LOG] => Начинаем обработку канала: {raw_channel_handle}")

    # 1) Получаем channelId через Selenium
    channel_id = get_channel_id_from_handle_selenium(raw_channel_handle)
    data["channel_id"] = channel_id

    # Создаём клиент YouTube Data API
    youtube = build("youtube", "v3", developerKey=DEVELOPER_KEY)

    if channel_id:
        try:
            # Запрашиваем данные о канале, включая statistics
            try:
                channel_response = youtube.channels().list(
                    part="snippet,brandingSettings,topicDetails,contentDetails,statistics",
                    id=channel_id
                ).execute()
            except HttpError as e:
                # Проверяем, не исчерпана ли квота
                if e.resp.status in [403, 429] or 'quotaExceeded' in str(e.content):
                    print("[LOG] -> Достигнут лимит квоты YouTube Data API (quotaExceeded). Прерываем обработку.")
                    raise
                else:
                    print(f"[LOG] -> HttpError при запросе channel API: {e}")
                    return data  # Возвращаем то, что есть

            ch_items = channel_response.get("items", [])
            if ch_items:
                channel_data = ch_items[0]

                # snippet
                snippet = channel_data.get("snippet", {})
                published_at = snippet.get("publishedAt", "")
                data["channel_creation_date_api"] = iso_to_readable(published_at)
                data["channel_country_api"] = snippet.get("country", "")

                # topicDetails
                topic_details = channel_data.get("topicDetails", {})
                topic_categories = topic_details.get("topicCategories", [])
                cleaned_topics = []
                for tcat in topic_categories:
                    if "wikipedia.org/wiki/" in tcat:
                        part = tcat.split("/wiki/")[-1].replace("_", " ")
                        cleaned_topics.append(part)
                    else:
                        cleaned_topics.append(tcat)
                data["channel_topics_api"] = ", ".join(cleaned_topics)

                # statistics
                stats = channel_data.get("statistics", {})
                view_count = stats.get("viewCount")
                data["total_views"] = int(view_count) if view_count else None

                # contentDetails (получаем uploads-плейлист)
                content_details = channel_data.get("contentDetails", {})
                rplaylists = content_details.get("relatedPlaylists", {})
                uploads_playlist_id = rplaylists.get("uploads", "")

                # Определяем даты самого нового/старого видео
                newest_date, oldest_date = get_newest_and_oldest_video_date_in_playlist(
                    uploads_playlist_id, youtube
                )
                data["last_video_published_api"]  = iso_to_readable(newest_date)
                data["first_video_published_api"] = iso_to_readable(oldest_date)

                # Собираем все videoId из плейлиста uploads
                all_video_ids = []
                next_page_token = None
                while True:
                    try:
                        playlist_resp = youtube.playlistItems().list(
                            part="contentDetails",
                            playlistId=uploads_playlist_id,
                            maxResults=50,
                            pageToken=next_page_token
                        ).execute()
                    except HttpError as e:
                        # Проверяем, не исчерпана ли квота
                        if e.resp.status in [403, 429] or 'quotaExceeded' in str(e.content):
                            print("[LOG] -> Достигнут лимит квоты YouTube Data API (quotaExceeded). Прерываем обработку.")
                            raise
                        else:
                            print(f"[LOG] -> HttpError при получении playlistItems: {e}")
                            break

                    items_page = playlist_resp.get("items", [])
                    if not items_page:
                        break

                    for pli_item in items_page:
                        vid_id = pli_item["contentDetails"]["videoId"]
                        all_video_ids.append(vid_id)

                    next_page_token = playlist_resp.get("nextPageToken")
                    if not next_page_token:
                        break

                # Подсчитываем, сколько из них шорт (≤ 60 сек)
                short_count = 0
                for batch in chunked(all_video_ids, 50):
                    batch_str = ",".join(batch)
                    try:
                        videos_resp = youtube.videos().list(
                            part="contentDetails",
                            id=batch_str
                        ).execute()
                    except HttpError as e:
                        # Проверяем, не исчерпана ли квота
                        if e.resp.status in [403, 429] or 'quotaExceeded' in str(e.content):
                            print("[LOG] -> Достигнут лимит квоты YouTube Data API (quotaExceeded). Прерываем обработку.")
                            raise
                        else:
                            print(f"[LOG] -> HttpError при получении videos: {e}")
                            continue

                    for item in videos_resp.get("items", []):
                        cdetails = item.get("contentDetails", {})
                        dur_str = cdetails.get("duration", "")
                        if not dur_str:
                            # Если нет ключа "duration" или значение пустое,
                            # пропускаем это видео или можно считать 0 секунд.
                            continue

                        seconds = parse_duration_to_seconds(dur_str)
                        if seconds <= 60:
                            short_count += 1


                # total_videos = всё
                data["total_videos"] = len(all_video_ids)
                # num_shorts = количество коротких
                data["num_shorts"] = short_count
                # num_videos = общее количество минус количество шорт
                data["num_videos"] = data["total_videos"] - data["num_shorts"]

                # Суммарные лайки/комментарии для всех видео (включая шорты)
                if all_video_ids:
                    try:
                        tlikes, tcomms = sum_likes_comments_via_api(all_video_ids, DEVELOPER_KEY)
                        data["estimated_likes"] = tlikes
                        data["estimated_comments"] = tcomms
                    except HttpError:
                        # Если поймали quotaExceeded в sum_likes_comments_via_api, завершится в вызывающем коде
                        pass

        except HttpError:
            # Если где-то внутри произошла quotaExceeded, завершаем
            print("[LOG] -> Прекращаем обработку (quotaExceeded).")
            raise

    # 3) Сбор Selenium (подписчики, email, город/страна из About, и пр.)
    try:
        channel_url = normalize_channel_url(raw_channel_handle)
        print(f"[LOG] => Построенный URL канала: {channel_url}")
        driver.get(channel_url)
        time.sleep(3)

        # Закрываем куки-баннер (если появится)
        try:
            cookie_btn = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "button[aria-label^='Accept the use of cookies']"))
            )
            cookie_btn.click()
            print("[LOG] => Куки-баннер обнаружен и закрыт.")
            time.sleep(2)
        except:
            print("[LOG] => Куки-баннер не найден/не кликабелен.")

        # Название канала
        try:
            h1_elem = driver.find_element(By.CSS_SELECTOR, "h1.dynamic-text-view-model-wiz__h1 span")
            channel_name = h1_elem.text.strip()
            print(f"[LOG] => Название канала (Selenium): {channel_name}")
            data["channel_name"] = channel_name
            data["first_last_name"] = guess_name_surname(channel_name)
        except:
            print("[LOG] => Не удалось найти h1.dynamic-text-view-model-wiz__h1 span")

        # Подписчики
        try:
            subs_elem = driver.find_element(By.XPATH, "//span[contains(text(),'subscriber')]")
            subs_text = subs_elem.text.strip()
            data["num_subscribers"] = parse_subscribers_to_int(subs_text)
            print(f"[LOG] => Подписчики (Selenium): {subs_text} -> {data['num_subscribers']}")
        except:
            print("[LOG] => Не удалось найти количество подписчиков (Selenium).")

        # Вкладка ABOUT
        about_url = channel_url.split("?")[0].rstrip("/") + "/about?hl=en&gl=US"
        print("[LOG] => Переходим на вкладку ABOUT:", about_url)
        driver.get(about_url)
        time.sleep(2)

        # Email
        try:
            desc_elems = driver.find_elements(By.CSS_SELECTOR, "div#description-container, yt-formatted-string#description")
            big_text = ""
            for d in desc_elems:
                big_text += d.text + "\n"
            emails = extract_emails_from_text(big_text)
            if emails:
                data["email"] = emails[0]
                print("[LOG] => Найден email:", data["email"])
        except:
            print("[LOG] => Не удалось извлечь email.")

        # Местоположение (город/страна)
        city_country = get_city_country_from_about(driver)
        if city_country:
            data["city_country"] = city_country
            print("[LOG] => Местоположение (Selenium):", city_country)

        # Дата создания (Selenium) — (храним, но не выводим)
        try:
            dt_joined = driver.find_element(By.XPATH, "//yt-formatted-string[contains(text(),'Joined')]").text.strip()
            dt_joined = dt_joined.replace("Joined", "").strip()
            data["creation_date"] = dt_joined
        except:
            pass

        # Вкладка CHANNELS (подсчёт, на сколько каналов автор подписан)
        channels_url = channel_url.split("?")[0].rstrip("/") + "/channels?hl=en&gl=US"
        driver.get(channels_url)
        time.sleep(2)
        try:
            channels = driver.find_elements(By.CSS_SELECTOR, "ytd-grid-channel-renderer, ytd-channel-renderer")
            data["num_following_channels"] = len(channels)
        except:
            data["num_following_channels"] = 0

    except Exception as e:
        print(f"!!! Ошибка при обработке канала {raw_channel_handle}: {e}")

    print("[LOG] => Завершили обработку канала:", raw_channel_handle)
    print("-"*60)
    return data


def main():
    driver = get_webdriver()

    if not os.path.exists(XLSX_INPUT):
        print(f"Не найден входной файл {XLSX_INPUT}")
        return

    # Загружаем входную книгу
    wb_in = load_workbook(XLSX_INPUT)
    ws_in = wb_in.active

    # Ищем столбец "channel_handle" во входном файле
    header_in = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    try:
        channel_index = header_in.index("channel_handle")
    except ValueError:
        print("В файле не найден столбец 'channel_handle'. Завершение.")
        return

    # Подготовка выходной книги
    already_processed = set()  # множество обработанных каналов
    if os.path.exists(XLSX_OUTPUT):
        # Если файл уже существует - открываем и считываем "channel_handle_in_excel"
        wb_out = load_workbook(XLSX_OUTPUT)
        ws_out = wb_out.active

        # Предполагаем, что заголовки уже есть в первой строке
        header_out = [cell.value for cell in next(ws_out.iter_rows(min_row=1, max_row=1))]
        if "channel_handle_in_excel" not in header_out:
            print(f"[LOG] -> Не найден нужный столбец 'channel_handle_in_excel' в {XLSX_OUTPUT}. Завершение.")
            driver.quit()
            return

        # Находим индекс колонки channel_handle_in_excel
        try:
            handle_col_index = header_out.index("channel_handle_in_excel") + 1
        except ValueError:
            print(f"[LOG] -> Столбец 'channel_handle_in_excel' отсутствует в выходном файле. Завершение.")
            driver.quit()
            return

        # Считываем все значения из этой колонки, кроме заголовка
        for row in ws_out.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                # row[0] соответствует channel_handle_in_excel, так как это первая колонка в выходном
                # но мы должны убедиться, что handle_col_index=1 => row[0]
                # если заголовки расположены иначе, использовать соответствующий индекс
                already_processed.add(str(row[0]))
    else:
        # Если файла нет, создаём новый
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "Channels"

        # Формируем заголовки для выходного файла
        headers = [
            "channel_handle_in_excel",
            "channel_id",
            "channel_name",
            "first_last_name",
            "city_country",
            "email",
            "num_subscribers",
            "total_videos",
            "num_videos",
            "num_shorts",
            "total_views",
            "channel_creation_date_api",
            "channel_country_api",
            "channel_topics_api",
            "first_video_published_api",
            "last_video_published_api",
            "num_following_channels",
            "estimated_likes",
            "estimated_comments"
        ]
        ws_out.append(headers)
        # Не забываем сохранить сразу, чтобы файл появился на диске
        wb_out.save(XLSX_OUTPUT)

    count_processed = 0

    try:
        for i, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=1):
            if MAX_CHANNELS is not None and count_processed >= MAX_CHANNELS:
                print(f"[LOG] => Достигнут лимит: обработано {MAX_CHANNELS} каналов. Прерываем цикл.")
                break

            raw_channel_handle = row[channel_index]
            if not raw_channel_handle:
                continue

            # Если уже обрабатывали этот канал, пропускаем
            if raw_channel_handle in already_processed:
                print(f"[LOG] => Канал {raw_channel_handle} уже есть в {XLSX_OUTPUT}, пропускаем.")
                continue

            print(f"\n=== [{i}] Обрабатываем канал из Excel: {raw_channel_handle} ===")
            try:
                data = process_channel(driver, raw_channel_handle)
            except HttpError:
                # Если попали сюда, значит quotaExceeded (или иная критическая ошибка)
                print("[LOG] => Остановка цикла из-за исчерпания квоты или критической ошибки API.")
                break

            # Формируем строку для итоговой таблицы
            row_out = [
                raw_channel_handle,                # handle из Excel
                data["channel_id"],
                data["channel_name"],
                data["first_last_name"],
                data["city_country"],
                data["email"],
                data["num_subscribers"],
                data["total_videos"],
                data["num_videos"],
                data["num_shorts"],
                data["total_views"],
                data["channel_creation_date_api"],
                data["channel_country_api"],
                data["channel_topics_api"],
                data["first_video_published_api"],
                data["last_video_published_api"],
                data["num_following_channels"],
                data["estimated_likes"],
                data["estimated_comments"]
            ]

            # Добавляем эту строку в выходной Excel
            ws_out.append(row_out)
            wb_out.save(XLSX_OUTPUT)  # Сохраняем после каждой записи, чтобы не потерять

            # Добавляем в множество "уже обработанных", чтобы не дублировать
            already_processed.add(raw_channel_handle)

            count_processed += 1

    finally:
        # Закрываем драйвер в любом случае
        driver.quit()

    print("[i] Запись данных завершена.")
    print("Готово!")


if __name__ == "__main__":
    main()
